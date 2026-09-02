#!/usr/bin/env python3
"""Ingest the customer crosswalk (data/raw/Nielsen_crosswalk.xlsx).

The crosswalk ties one customer together across systems: the internal
hierarchy (Territory -> Parent Account -> Sales Account, with Team Lead and
Account Lead), the Nielsen trading area (NIQ Match), and — matched here by
normalized name — the Telus promotion customer ids.

Emits lib/fixtures/crosswalk.json (one row per sales account x NIQ match) and
prints a match report. Re-run after replacing the raw workbook.
"""
import gzip
import json
import pathlib
import re
import sys

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw" / "Nielsen_crosswalk.xlsx"
FIXTURES = ROOT / "lib" / "fixtures"

# NIQ TA name -> the app's market_code (matches scripts/ingest_albsco.py).
TA_TO_MARKET = {
    "ALBSCO Acme TA": "ALB-ACME",
    "ALBSCO Denver Div TA": "ALB-DENVER",
    "ALBSCO Eastern TA": "ALB-EASTERN",
    "ALBSCO Intermountain Div TA": "ALB-INTMTN",
    "ALBSCO Jewel Div TA": "ALB-JEWEL",
    "ALBSCO Nor Cal Div TA": "ALB-NORCAL",
    "ALBSCO Portland Div TA": "ALB-PORTLAND",
    "ALBSCO Seattle Div TA": "ALB-SEATTLE",
    "ALBSCO Shaws Div TA": "ALB-SHAWS",
    "ALBSCO So Cal Div TA": "ALB-VONS",
    "ALBSCO Southern Div TA": "ALB-SOUTHERN",
    "ALBSCO Southwest Div TA": "ALB-SOUTHWEST",
    "ALBSCO United Div TA": "ALB-UNITED",
}

CODE_TOKEN = re.compile(r"^[A-Z]{2,5}\d{2,4}[A-Z]?$")

def norm_name(s: str) -> str:
    """Normalize a customer name for cross-system matching: drop parentheticals,
    Telus account-code tokens (PUB100, WAL100...), punctuation and case."""
    s = re.sub(r"\([^)]*\)", " ", s)
    tokens = [t for t in re.split(r"[\s,]+", s) if t and not CODE_TOKEN.match(t)]
    return re.sub(r"[^a-z0-9]", "", " ".join(tokens).lower())

def main() -> None:
    wb = load_workbook(RAW, read_only=True, data_only=True)
    ws = wb["Reporting Tab"]
    raw_rows = [r for r in ws.iter_rows(min_row=7, values_only=True) if any(v is not None for v in r)]

    # Telus customers, for the name join
    promos = json.load(gzip.open(ROOT / "data" / "promos" / "promotions.json.gz", "rt"))
    telus = {}
    for p in promos:
        telus.setdefault(norm_name(p["customer_name"]), set()).add((p["customer_id"], p["customer_name"]))

    rows, flags = [], []
    seen_ids = set()
    seen_nat = set()  # exact-duplicate guard: (customer_name, sales_account, niq_match)
    for r in raw_rows:
        nm, cls, terr, parent, sales, tlead, alead, niq = (str(v).strip() if v is not None else None for v in r[:8])
        if not nm:
            continue
        nat = (nm, sales, niq)
        if nat in seen_nat:
            flags.append(f"duplicate row dropped: {nm!r} -> {niq!r}")
            continue
        seen_nat.add(nat)
        if terr and "?" in terr:
            flags.append(f"data question left in sheet: {nm!r} territory={terr!r}")
        market = TA_TO_MARKET.get(niq) if niq else None
        if niq and market is None and niq not in TA_TO_MARKET:
            market = None  # non-ALBSCO TA — named market we hold no data for yet
        t = telus.get(norm_name(nm), set())
        row_id = re.sub(r"[^a-z0-9]+", "-", nm.lower()).strip("-")
        if market:
            row_id += "-" + market.lower()
        if row_id in seen_ids:
            row_id += "-" + str(sum(1 for i in seen_ids if i.startswith(row_id)))
        seen_ids.add(row_id)
        rows.append({
            "id": row_id,
            "customer_name": nm,
            "customer_class": cls,              # RT retail | CL club
            "territory": terr,
            "parent_account": parent,
            "sales_account": sales,
            "team_lead": tlead,
            "account_lead": alead,
            "niq_match": niq,                    # NIQ TA name, or null
            "market_code": market,               # app market code when we hold that TA's data
            "telus_customer_ids": sorted(i for i, _ in t),
            "telus_customer_names": sorted(n for _, n in t),
        })

    FIXTURES.mkdir(parents=True, exist_ok=True)
    (FIXTURES / "crosswalk.json").write_text(json.dumps(rows, indent=1) + "\n")

    matched_telus = sum(1 for x in rows if x["telus_customer_ids"])
    with_niq = sum(1 for x in rows if x["niq_match"])
    with_data = sum(1 for x in rows if x["market_code"])
    print(f"crosswalk.json: {len(rows)} rows")
    print(f"  NIQ match named: {with_niq} · with Nielsen data on file: {with_data}")
    print(f"  Telus customers matched by name: {matched_telus}")
    for f in flags:
        print("  ⚠", f)
    # Telus customers that matched nothing (for the report)
    matched_ids = {i for x in rows for i in x["telus_customer_ids"]}
    unmatched = sorted({(p["customer_id"], p["customer_name"]) for p in promos if p["customer_id"] not in matched_ids})
    print(f"  Telus customers with no crosswalk row: {len(unmatched)} (first 12)")
    for i, n in unmatched[:12]:
        print("    ", i, n)

if __name__ == "__main__":
    sys.exit(main())
