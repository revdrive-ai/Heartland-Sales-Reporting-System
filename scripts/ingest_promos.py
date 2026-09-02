#!/usr/bin/env python3
"""Ingest the Telus promotions import template (data/raw/*.xlsx) into the app.

Source: Heartland_Promo_Import_Template.xlsx — the transformed Telus retail
promotions export (snapshot 2026-08-21, FY2026). Two linked tables joined on
promo_id, per the workbook's own Import Spec:

  data/promos/promotions.json.gz   one row per promotion (header)
  data/promos/promo-lines.json.gz  one row per component + item (the money)
  data/promos/meta.json            snapshot date, counts, reconciled totals
  lib/fixtures/promo-enums.json    the Valid Values tab (controlled vocab)

Validation follows the spec sheet: unique keys, end_date >= start_date, enums
must match Valid Values, amounts numeric. line_count / planned / actual on
each promotion are recomputed from the lines (the sheet's own columns are
live-formula checks, not payload) and drift is reported.

Run:  python3 scripts/ingest_promos.py
"""
import gzip
import json
import pathlib
import sys
from collections import defaultdict

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw" / "Heartland_Promo_Import_Template.xlsx"
OUT = ROOT / "data" / "promos"
FIXTURES = ROOT / "lib" / "fixtures"
SNAPSHOT_DATE = "2026-08-21"  # per the Import Spec sheet (export of 08/21/2026)

def rows_of(ws):
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(it)]
    for raw in it:
        if all(v is None for v in raw):
            continue
        yield dict(zip(header, raw))

def iso(v):
    s = str(v).strip()
    return s[:10]

def money(v):
    if v is None or str(v).strip() == "":
        return 0.0
    return round(float(str(v).replace("$", "").replace(",", "")), 2)

def main() -> None:
    wb = load_workbook(RAW, read_only=True, data_only=True)
    errors = []

    # ---- controlled vocabulary ----------------------------------------------
    vv = list(wb["Valid Values"].iter_rows(values_only=True))
    enums = {}
    for c, h in enumerate(vv[0]):
        enums[str(h)] = [str(r[c]) for r in vv[1:] if r[c] is not None]
    FIXTURES.mkdir(parents=True, exist_ok=True)
    (FIXTURES / "promo-enums.json").write_text(json.dumps(enums, indent=1) + "\n")
    print("promo-enums.json:", {k: len(v) for k, v in enums.items()})

    # ---- promotions ----------------------------------------------------------
    promos = []
    seen = set()
    for r in rows_of(wb["Promotions"]):
        pid = str(r["promo_id"]).strip()
        if pid in seen:
            errors.append(f"duplicate promo_id {pid}")
            continue
        seen.add(pid)
        start, end = iso(r["start_date"]), iso(r["end_date"])
        if end < start:
            errors.append(f"{pid}: end_date {end} before start_date {start}")
        for field, enum in [("promo_status", "promo_status"), ("template_type", "template_type"),
                            ("performance_type", "performance_type"), ("channel", "channel"),
                            ("market", "market")]:
            if str(r[field]) not in enums[enum]:
                errors.append(f"{pid}: {field} {r[field]!r} not in Valid Values")
        promos.append({
            "promo_id": pid,
            "promo_title": " ".join(str(r["promo_title"]).split()),
            "fiscal_year": int(r["fiscal_year"]),
            "promo_status": str(r["promo_status"]),
            "template_type": str(r["template_type"]),
            "performance_type": str(r["performance_type"]),
            "customer_id": str(r["customer_id"]),
            "customer_name": str(r["customer_name"]).strip(),
            "customer_code": (str(r["customer_code"]).strip() or None) if r["customer_code"] is not None else None,
            "channel": str(r["channel"]),
            "market": str(r["market"]),
            "planner_template": str(r["planner_template"]).strip(),
            "start_date": start,
            "end_date": end,
            # line_count / planned_amount / actual_amount are recomputed below
        })

    # ---- promo lines ---------------------------------------------------------
    lines = []
    seen_lines = set()
    promo_ids = {p["promo_id"] for p in promos}
    for r in rows_of(wb["Promo_Lines"]):
        lid = str(r["line_id"]).strip()
        if lid in seen_lines:
            errors.append(f"duplicate line_id {lid}")
            continue
        seen_lines.add(lid)
        pid = str(r["promo_id"]).strip()
        if pid not in promo_ids:
            errors.append(f"line {lid}: unknown promo_id {pid}")
        if str(r["rate_uom"]) not in enums["rate_uom"]:
            errors.append(f"line {lid}: rate_uom {r['rate_uom']!r} not in Valid Values")
        if str(r["component_type"]) not in enums["component_type"]:
            errors.append(f"line {lid}: component_type {r['component_type']!r} not in Valid Values")
        lines.append({
            "line_id": lid,
            "promo_id": pid,
            "component_type": str(r["component_type"]),
            "brand": str(r["brand"]),
            "item_number": str(r["item_number"]).strip(),
            "item_description": str(r["item_description"]).strip() if r["item_description"] is not None else None,
            "rate": round(float(r["rate"] or 0), 4),
            "rate_uom": str(r["rate_uom"]),
            "planned_amount": money(r["planned_amount"]),
            "actual_amount": money(r["actual_amount"]),
        })

    # ---- rollups (the sheet's shaded columns are checks; we recompute) -------
    agg = defaultdict(lambda: {"n": 0, "p": 0.0, "a": 0.0})
    for l in lines:
        a = agg[l["promo_id"]]
        a["n"] += 1
        a["p"] += l["planned_amount"]
        a["a"] += l["actual_amount"]
    for p in promos:
        a = agg[p["promo_id"]]
        p["line_count"] = a["n"]
        p["planned_amount"] = round(a["p"], 2)
        p["actual_amount"] = round(a["a"], 2)
    orphan_promos = [p["promo_id"] for p in promos if p["line_count"] == 0]

    if errors:
        print(f"\n{len(errors)} VALIDATION ERRORS:")
        for e in errors[:25]:
            print(" •", e)
        raise SystemExit(1)

    OUT.mkdir(parents=True, exist_ok=True)
    with gzip.open(OUT / "promotions.json.gz", "wt", encoding="utf-8") as f:
        json.dump(promos, f, separators=(",", ":"))
    with gzip.open(OUT / "promo-lines.json.gz", "wt", encoding="utf-8") as f:
        json.dump(lines, f, separators=(",", ":"))

    meta = {
        "source_file": RAW.name,
        "snapshot_date": SNAPSHOT_DATE,
        "fiscal_year": 2026,
        "promotions": len(promos),
        "promo_lines": len(lines),
        "planned_total": round(sum(l["planned_amount"] for l in lines), 2),
        "actual_total": round(sum(l["actual_amount"] for l in lines), 2),
        "promos_without_lines": len(orphan_promos),
    }
    (OUT / "meta.json").write_text(json.dumps(meta, indent=1) + "\n")

    print(f"promotions.json.gz: {len(promos)} promos")
    print(f"promo-lines.json.gz: {len(lines)} lines")
    print(f"planned ${meta['planned_total']:,.2f} · actual ${meta['actual_total']:,.2f}")
    print(f"promos without lines: {len(orphan_promos)}")
    by_status = defaultdict(int)
    for p in promos:
        by_status[p["promo_status"]] += 1
    print("by status:", dict(by_status))

if __name__ == "__main__":
    sys.exit(main())
